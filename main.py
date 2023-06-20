# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Alignment
from ProcessExcel import process_array
import numpy as np
from array import *



def print_array(input_array):
    for row in input_array:
        print(row)


def reformat_sheet(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Set light blue color to the first row
    first_row = sheet[1]
    for cell in first_row:
        cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")

    # Center all text in cells
    for row in sheet.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths to fit cell contents
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_path)


def update_input_sheet(file_path, sheet_name, array):
    wb = openpyxl.load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        wb.create_sheet(title=sheet_name)
        sheet = wb[sheet_name]
        
    # Clear previous data in the sheet
    sheet.delete_rows(1, sheet.max_row)

    # Add the array values to the sheet
    for row in array:
        sheet.append(row)
    
    reformat_sheet(file_path, sheet_name)
    # Save the workbook
    wb.save(file_path)
    

def main():
    input_file_path = "workbook1.xlsx"
    default_sheet_name = "Sheet1"
    input_sheet_name = "Input Sheet"

    input_columns_to_check = [2, 4, 5, 7]  # Specify the columns to concatenate (e.g., A=1, B=2, D=4)

    ProcessedState_InputSheet, ProcessedState_UniqueIndices = process_array(input_file_path, default_sheet_name, input_columns_to_check)
    update_input_sheet(input_file_path, input_sheet_name, ProcessedState_InputSheet)


if __name__ == "__main__":
    main()


