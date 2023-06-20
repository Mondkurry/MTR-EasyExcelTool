import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import numpy as np

def print_array(input_array):
  	for row in input_array:
    		print(row)
                

def populate_array_from_excel(file_path, sheet_name):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Create an empty array
    result_array = []

    # Iterate over the rows and columns and append values to the array
    for row in sheet.iter_rows(values_only=True):
        result_array.append(list(row))

    return result_array


def extract_columns(input_array, check_columns):
    output_array = []
    
    for row in input_array:
        output_row = [row[index] for index in check_columns]
        output_array.append(output_row)

    return output_array

def compare_arrays(array1, array2): # Compares two arrays and returns the values for indices in array 1 that are present in array2
    indexes = []
    for value in array2:
        if value in array1:
            index = array1.index(value)
            indexes.append(index)
    return indexes


def get_unique_rows(input_array, check_columns):

    Array_OnlyCheckColumns = extract_columns(input_array, check_columns)

    unique_rows_onlyCheckColumns = []
    unique_rows = []
    index_of_unique = []
    index = 0

    for row in Array_OnlyCheckColumns:
        if row not in unique_rows_onlyCheckColumns:
            # index_of_unique.append(index)
            unique_rows_onlyCheckColumns.append(row)
            unique_rows.append(input_array[index])
        index += 1

    index_of_unique = compare_arrays(unique_rows_onlyCheckColumns, Array_OnlyCheckColumns)
    
    return index_of_unique, unique_rows


def process_array(file_path, sheet_name, columns_to_check):
    DefaultState_Sheet = populate_array_from_excel(file_path, sheet_name)
    ProcessedState_UniqueIndices, ProcessedState_Sheet = get_unique_rows(DefaultState_Sheet, columns_to_check)
    
    return ProcessedState_Sheet, ProcessedState_UniqueIndices

def update_input_sheet(file_path, sheet_name, array):
    wb = openpyxl.load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
        print(1)
    else:
        sheet = wb.create_sheet(title=sheet_name)
        print(2)

    # Add the array values to the sheet
    for row in array:
        sheet.append(row)
    
        # Save the workbook
    wb.save(file_path)



def reformat_excel_sheet(file_path, sheet_name):
    # Load the workbook
    workbook = load_workbook(file_path)
    sheet2 = workbook[sheet_name]

    # Apply light blue color to the first row
    light_blue_fill = PatternFill(start_color='B7C9E3', end_color='B7C9E3', fill_type='solid')
    for cell in sheet2[1]:
        cell.fill = light_blue_fill

    # Center-align all cells
    for row in sheet2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Auto-fit column width to contain text
    for column in sheet2.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add some padding
        sheet2.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the modified workbook
    workbook.save(file_path)


def process_and_update_userinput_sheet(file_path, default_sheet_name, userinput_sheet_name, columns_to_check):
    ProcessedState_InputSheet, ProcessedState_UniqueIndices = process_array(
        file_path, default_sheet_name, columns_to_check)
    
    update_input_sheet(file_path, userinput_sheet_name,
                       ProcessedState_InputSheet)
    
    reformat_excel_sheet(file_path, userinput_sheet_name)