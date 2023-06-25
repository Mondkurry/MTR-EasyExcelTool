import openpyxl
from openpyxl import load_workbook
from ProcessExcel import populate_array_from_excel, print_array
from UpdateInputSheet import reformat_excel_sheet, process_array

def populate_output(file_path, default_sheet, input_sheet, output_sheet, processedState_sheet, processedState_uniqueIndices):
    wb = openpyxl.load_workbook(file_path)
    default_sheet = wb[default_sheet]
    input_sheet = wb[input_sheet]
    
    if output_sheet in wb.sheetnames:
        sheet = wb[output_sheet]
        sheet.delete_rows(1, sheet.max_row)
        print(f"Writing to existing Sheet: {output_sheet}")
    else:
        sheet = wb.create_sheet(title=output_sheet)
        print(f"Creating new Sheet: {output_sheet}")
    
    for index in processedState_uniqueIndices:
        if index != len(processedState_sheet):
            sheet.append(processedState_sheet[index])
    
    # replace the first two columns of sheet with the first two columns of default_sheet    
    for row in range(1, sheet.max_row + 1):
        for column in range(1, 3):
            sheet.cell(row=row, column=column).value = default_sheet.cell(row=row, column=column).value

    wb.save(file_path)  


def populate_output_andReformat(file_path, default_sheet_name, input_sheet, output_sheet, columns_to_check):
    ProcessedState_InputSheet, ProcessedState_UniqueIndices = process_array(file_path, default_sheet_name, columns_to_check)
    processed_input_sheet = populate_array_from_excel(file_path, input_sheet)

    populate_output(file_path, default_sheet_name, input_sheet, output_sheet, processed_input_sheet, ProcessedState_UniqueIndices)
    reformat_excel_sheet(file_path, output_sheet)