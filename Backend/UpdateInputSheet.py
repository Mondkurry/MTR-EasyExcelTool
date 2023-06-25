import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from ProcessExcel import process_array

def update_input_sheet(file_path, sheet_name, array):
    wb = openpyxl.load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
        print(f"Writing to existing Sheet: {sheet_name}")
    else:
        sheet = wb.create_sheet(title=sheet_name)
        print(f"Creating new Sheet: {sheet_name}")

    # Add the array values to the sheet
    for row in array:
        sheet.append(row)
    
        # Save the workbook
    wb.save(file_path)  



def reformat_excel_sheet(file_path, sheet_name):
    # Load the workbook
    workbook = load_workbook(file_path)
    sheet2 = workbook[sheet_name]

    # Apply light blue color to the first row
    light_blue_fill = PatternFill(start_color='B7C9E3', end_color='B7C9E3', fill_type='solid')
    for cell in sheet2[1]:
        cell.fill = light_blue_fill

    # Center-align all cells
    for row in sheet2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Auto-fit column width to contain text
    for column in sheet2.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add some padding
        sheet2.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the modified workbook
    workbook.save(file_path)


def process_and_update_userinput_sheet(file_path, default_sheet_name, input_sheet_name, columns_to_check):
    ProcessedState_InputSheet, ProcessedState_UniqueIndices = process_array(
        file_path, default_sheet_name, columns_to_check)
    
    update_input_sheet(file_path, input_sheet_name,
                       ProcessedState_InputSheet)
    
    reformat_excel_sheet(file_path, input_sheet_name)